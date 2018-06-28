import keras
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.metrics import confusion_matrix
import pickle
# Own Scripts
from gold_package.features_labels_setup import Features_labels
from gold_package.SNN_setup import create_SNN_hparams, SNN, Siamese_loader


def read_reaction_data(print_mode=False):
    df = pd.read_excel('./excel/reaction_data_template.xlsx', sheet_name='df')
    features_c = pd.read_excel('./excel/reaction_data_template.xlsx', sheet_name='features_c').values
    features_d = pd.read_excel('./excel/reaction_data_template.xlsx', sheet_name='features_d').values
    # There should be one column in the excel sheet for labels only!
    labels = pd.read_excel('./excel/reaction_data_template.xlsx', sheet_name='labels').values.flatten()
    Features_labels(features_c, features_d, labels, save_mode=True)
    if print_mode:
        print(df)
        print('Continuous features : \n {}'.format(features_c))
        print('Discrete features : \n {}'.format(features_d))
        print('Labels : \n {}'.format(labels))
    return features_c, features_d, labels, df


def train(fl):
    hparams = create_SNN_hparams(reg_term=2e-4)
    model = SNN(hparams, fl.features_c_dim, fl.features_d_hot_dim)
    loader = Siamese_loader(model.siamese_net, fl)
    loader.train(steps=100, batch_size=64, save_mode=True)


def generate_new_results(fl, loader, numel, save_to_excel=False):
    def print_array_to_excel(array, first_cell, ws):
        shape = array.shape
        if shape[1] == 1:
            for i in range(shape[0]):
                j = 0
                ws.cell(i + first_cell[0], j + first_cell[1]).value = array[i][0]
        else:
            for i in range(shape[0]):
                for j in range(shape[1]):
                    ws.cell(i + first_cell[0], j + first_cell[1]).value = array[i, j]

    gen_dic = fl.generate_random_examples(numel)
    predicted_labels = loader.oneshot_predict(gen_dic['gen_features_c_norm_a'], gen_dic['gen_features_d_hot_a'])
    if save_to_excel:
        wb = load_workbook('./excel/reaction_data_template.xlsx')
        ws_gen_features_c = wb['gen_features_c']
        ws_gen_features_d = wb['gen_features_d']
        ws_gen_predicted_labels = wb['gen_predicted_labels']
        print_array_to_excel(gen_dic['gen_features_c_a'], [2, 1], ws_gen_features_c)
        # [:,None] is to change 1d ndarray to a nx1 ndarry.
        # But for some reason, when indexing row vector, must be [i][0] with the extra [0]
        print_array_to_excel(gen_dic['gen_features_d_dense_a'][:, None], [2, 1], ws_gen_features_d)
        print_array_to_excel(np.array(predicted_labels)[:, None], [2, 1], ws_gen_predicted_labels)
        wb.save('./excel/reaction_data_temp.xlsx')
    return predicted_labels


read_reaction_data(False)
fl = pickle.load(open('./save/features_labels/fl.obj', 'rb'))
train(fl)
model = keras.models.load_model('./save/models/SNN.h5')
generate_new_results(fl, Siamese_loader(model, fl), 5000, save_to_excel=True)
