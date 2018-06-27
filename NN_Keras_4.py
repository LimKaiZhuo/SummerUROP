import keras as K
from keras.models import Sequential
from keras.layers import Dense, Dropout
from keras import regularizers, optimizers
from keras.callbacks import EarlyStopping, ModelCheckpoint
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.metrics import confusion_matrix
import os, itertools, pickle
# Own Scripts
from UROP_Package.features_labels_setup import Features, Labels, feature_splitter
from UROP_Package.reactions_list import denbigh_rxn, denbigh_rxn2, create_class
from UROP_Package.DNN_setup import create_random_features, create_hparams, New_DNN, DNN_classifer
from UROP_Package.GAN_setup import create_GAN_hparams, create_normalized_feature_label, \
    save_new_training_testing_validation_examples, GAN, DAGAN, generating_new_examples_using_GAN, \
    generating_new_examples_using_DAGAN, \
    compare_GAN_to_ODE
from UROP_Package.SNN_setup import SNN, Siamese_loader, create_SNN_hparams


def SNN_prediction(training_features, validation_features, bins=[10, 6, 0], load_model=False):
    """

    :param training_features: Feature class
    :param validation_features: Feature class
    :param load_model: Whether to load an existing model in ./save/models/SNN.h5
    :return: Confusion matrix
    """

    x_train = create_class(training_features, denbigh_rxn2, bins, print_mode=True)
    if load_model:
        model = K.models.load_model('./save/models/SNN.h5')
        data = Siamese_loader(model, x_train, x_train, bins=bins)
    else:
        SNN_hparams = create_SNN_hparams()
        model = SNN(SNN_hparams, features_dim=3).siamese_net
        data = Siamese_loader(model, x_train, x_train,bins=bins)
        data.train(steps=100, batch_size=64, save_mode=True)
    _, cm, acc = data.oneshot_predict(validation_features.n_features, compare_mode=True)
    return cm, acc


def DNN_prediction(training_features, validation_features, reaction=denbigh_rxn2, bins=[10, 6, 0], load_model=False):
    if load_model:
        model = K.models.load_model('./save/models/DNN_training_only.h5')
    else:
        hparams = create_hparams(epochs=100, batch_size=64)
        DNN = New_DNN(3, 5, hparams)
        model = DNN.train_model_epochs_training_data_only(training_features, reaction(training_features),
                                                          save_mode=True)
    validation_labels = reaction(validation_features)
    actual_class_store = np.digitize(validation_labels.a_labels[:, 1], bins) - 1
    predicted_validation_labels = Labels(model.predict(validation_features.n_features), mode='n')
    predicted_class_store = np.digitize(predicted_validation_labels.a_labels[:, 1], bins) - 1
    cm = confusion_matrix(actual_class_store, predicted_class_store)
    accuracy = np.count_nonzero(
        np.array(actual_class_store) - np.array(predicted_class_store) == 0) / validation_features.n_features.shape[
                   0] * 100
    return cm, accuracy


def cDNN_prediction(training_features, validation_features, reaction=denbigh_rxn2, bins=[10, 6, 0], load_model=False):
    if load_model:
        model = K.models.load_model('./save/models/cDNN_training_only.h5')
    else:
        hparams = create_hparams(epochs=100, batch_size=64)
        DNN = DNN_classifer(3, hparams)
        model = DNN.train_model_epochs_training_data_only(training_features, reaction(training_features),
                                                          save_mode=True)
    actual_class_store = reaction(validation_features).binning(bins)
    predicted_class_store = np.ndarray.flatten(np.around(model.predict(validation_features.n_features)))
    cm = confusion_matrix(actual_class_store, predicted_class_store)
    accuracy = np.count_nonzero(
        np.array(actual_class_store) - np.array(predicted_class_store) == 0) / validation_features.n_features.shape[
                   0] * 100
    return cm, accuracy


def print_array_to_excel(array,first_cell,ws):
    shape=array.shape
    for i in range(shape[0]):
        for j in range(shape[1]):
            ws.cell(i+first_cell[0],j+first_cell[1]).value=array[i,j]

# create_random_features(20,1,'training_features',save_mode=True)
# create_random_features(1000,1,'prediction_features',save_mode=True)

def comparing_SNN_to_DNN(bins=[10, 6, 0], load_model=[True,True,True]):
    wb=load_workbook('./Project Excel/SNN.xlsx')
    bins_store=[[10,6,0],[10,6,4,0],[10,6,4,2,0]]
    for run in range(3):
        bins=bins_store[run]
        ws=wb.get_sheet_by_name(str(run))
        for cnt in range(3):
            training_features = pickle.load(open('./save/features_labels/training_features.obj', 'rb'))
            validation_features = pickle.load(open('./save/features_labels/validation_features.obj', 'rb'))
            cm_DNN, acc_DNN = DNN_prediction(training_features, validation_features, bins=bins, load_model=load_model[0])
            cm_cDNN, acc_cDNN = cDNN_prediction(training_features, validation_features, bins=bins, load_model=load_model[1])
            cm_SNN, acc_SNN = SNN_prediction(training_features, validation_features, bins=bins, load_model=load_model[2])
            print('cm_DNN = \n{}'.format(cm_DNN))
            print('cm_cDNN = \n{}'.format(cm_cDNN))
            print('cm_SNN = \n{}'.format(cm_SNN))
            print('acc_DNN = {} , acc_cDNN = {} , acc_SNN = {}'.format(acc_DNN, acc_cDNN, acc_SNN))
            ws.cell(cnt+2,3).value=acc_DNN
            ws.cell(cnt + 2, 4).value = acc_cDNN
            ws.cell(cnt + 2, 5).value = acc_SNN
            print_array_to_excel(cm_DNN,[cnt*4+7,3],ws)
            print_array_to_excel(cm_cDNN, [cnt * 4 + 7, 7], ws)
            print_array_to_excel(cm_SNN, [cnt * 4 + 7, 11], ws)
    wb.save('./Project Excel/SNN_temp.xlsx')


comparing_SNN_to_DNN([10,6,3,0],load_model=[False,False,False])
